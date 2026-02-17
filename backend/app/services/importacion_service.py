"""
Servicio de importación masiva desde Excel.

Estrategia de performance:
- openpyxl en modo read_only para bajo consumo de memoria
- Validación en memoria (set de duplicados, checks por fila)
- INSERT por lotes de 100 filas con flush intermedio → evita timeout de 30s
- Un solo commit al final si todo OK
"""

from io import BytesIO
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, insert
from fastapi import HTTPException, status

from app.models.tenant import Cliente, Producto
import logging

logger = logging.getLogger(__name__)

# ============================================================================
# CONSTANTES
# ============================================================================

BATCH_SIZE = 100  # Filas por flush — equilibrio entre velocidad y seguridad

# ── Estilos de la plantilla ──
_HEADER_FILL_REQ = PatternFill(start_color="FF6D28", end_color="FF6D28", fill_type="solid")
_HEADER_FILL_OPT = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FONT_OPT = Font(bold=True, color="333333", size=11)
_HELP_FONT = Font(italic=True, color="888888", size=9)
_HELP_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _apply_header_style(ws, col: int, is_required: bool):
    """Aplica estilo al header (fila 1) y help (fila 2)."""
    cell = ws.cell(row=1, column=col)
    cell.fill = _HEADER_FILL_REQ if is_required else _HEADER_FILL_OPT
    cell.font = _HEADER_FONT if is_required else _HEADER_FONT_OPT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _THIN_BORDER

    help_cell = ws.cell(row=2, column=col)
    help_cell.font = _HELP_FONT
    help_cell.fill = _HELP_FILL
    help_cell.alignment = Alignment(horizontal="center", wrap_text=True)
    help_cell.border = _THIN_BORDER


# ============================================================================
# PLANTILLAS
# ============================================================================

# (header, help_text, is_required, example_value)
CLIENTE_COLUMNS = [
    ("tipo_documento*", "RUC | DNI | CE | PASAPORTE", True, "RUC"),
    ("numero_documento*", "Nro. sin guiones (ej: 20100947020)", True, "20100947020"),
    ("razon_social*", "Nombre completo o razón social", True, "Empresa ABC S.A.C."),
    ("nombre_comercial", "Nombre comercial (opcional)", False, "ABC"),
    ("email", "correo@ejemplo.com", False, "contacto@abc.com"),
    ("telefono", "(01) 456-7890", False, "01-4567890"),
    ("direccion_completa", "Dirección fiscal completa", False, "Av. Javier Prado 123, San Isidro"),
    ("ubigeo", "Código SUNAT 6 dígitos (ej: 150101)", False, "150101"),
]

PRODUCTO_COLUMNS = [
    ("codigo*", "SKU único del producto", True, "PROD-001"),
    ("nombre*", "Nombre del producto o servicio", True, "Laptop HP 15"),
    ("precio_unitario*", "Precio sin IGV (número decimal)", True, 2500.00),
    ("tipo", "producto | servicio | combo (default: producto)", False, "producto"),
    ("descripcion", "Descripción breve", False, "Laptop con 16GB RAM y SSD 512GB"),
    ("categoria", "Categoría del producto", False, "Electrónica"),
    ("marca", "Marca", False, "HP"),
    ("costo_unitario", "Costo de compra (número decimal)", False, 1800.00),
    ("moneda", "PEN | USD (default: PEN)", False, "PEN"),
    ("unidad_medida", "UND | KG | LTR | DOC | etc (default: UND)", False, "UND"),
    ("igv_porcentaje", "Porcentaje IGV (default: 18)", False, 18),
]


def _build_template(columns: list, sheet_name: str) -> BytesIO:
    """Genera un archivo Excel con headers estilizados y fila de ejemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, (header, help_text, is_required, example) in enumerate(columns, 1):
        # Header row
        ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(ws, col_idx, is_required)

        # Help row
        ws.cell(row=2, column=col_idx, value=help_text)

        # Example row
        ws.cell(row=3, column=col_idx, value=example)

        # Auto-width
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            len(str(header)) + 4, len(str(help_text)) + 2, 18
        )

    # Freeze header rows
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generar_plantilla_clientes() -> BytesIO:
    return _build_template(CLIENTE_COLUMNS, "Clientes")


def generar_plantilla_productos() -> BytesIO:
    return _build_template(PRODUCTO_COLUMNS, "Productos")


# ============================================================================
# IMPORTACIÓN
# ============================================================================

def _clean(val: Any) -> str | None:
    """Limpia un valor de celda: strips strings, convierte None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


async def importar_clientes(db: AsyncSession, file_bytes: bytes) -> dict:
    """
    Importa clientes desde un archivo Excel.
    
    Performance:
    - Lee el archivo en modo read_only
    - Valida todo en memoria primero (O(n) con set para duplicados)
    - INSERT por lotes de BATCH_SIZE con flush intermedio
    - Verifica duplicados contra la BD con una sola query
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Archivo Excel inválido o corrupto")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # Skip header + help rows
    wb.close()

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene datos (las filas empiezan en la fila 3)")

    # ── Fase 1: Validación en memoria ──
    errores = []
    registros_validos = []
    docs_en_archivo = set()
    tipos_validos = {"RUC", "DNI", "CE", "PASAPORTE"}

    for i, row in enumerate(rows, start=3):
        # Ensure row has enough columns
        padded = list(row) + [None] * (8 - len(row)) if len(row) < 8 else list(row)

        tipo_doc = _clean(padded[0])
        num_doc = _clean(padded[1])
        razon = _clean(padded[2])

        # Validar campos obligatorios
        if not tipo_doc or not num_doc or not razon:
            errores.append({"fila": i, "error": "Faltan campos obligatorios (tipo_documento, numero_documento, razon_social)"})
            continue

        tipo_doc = tipo_doc.upper()
        if tipo_doc not in tipos_validos:
            errores.append({"fila": i, "error": f"tipo_documento inválido: '{tipo_doc}'. Use: RUC, DNI, CE o PASAPORTE"})
            continue

        # Duplicados dentro del archivo
        if num_doc in docs_en_archivo:
            errores.append({"fila": i, "error": f"numero_documento '{num_doc}' duplicado en el archivo"})
            continue
        docs_en_archivo.add(num_doc)

        registros_validos.append({
            "tipo_documento": tipo_doc,
            "numero_documento": num_doc,
            "razon_social": razon,
            "nombre_comercial": _clean(padded[3]),
            "email": _clean(padded[4]),
            "telefono": _clean(padded[5]),
            "direccion_completa": _clean(padded[6]),
            "ubigeo": _clean(padded[7]),
            "estado": "activo",
        })

    if not registros_validos:
        return {"creados": 0, "errores": errores}

    # ── Fase 2: Verificar duplicados contra la BD (una sola query) ──
    docs_a_insertar = [r["numero_documento"] for r in registros_validos]
    result = await db.execute(
        select(Cliente.numero_documento).where(
            Cliente.numero_documento.in_(docs_a_insertar),
            Cliente.deleted_at == None,
        )
    )
    docs_existentes = set(result.scalars().all())

    registros_finales = []
    for r in registros_validos:
        if r["numero_documento"] in docs_existentes:
            errores.append({"fila": "—", "error": f"numero_documento '{r['numero_documento']}' ya existe en la base de datos"})
        else:
            registros_finales.append(r)

    if not registros_finales:
        return {"creados": 0, "errores": errores}

    # ── Fase 3: INSERT por lotes ──
    creados = 0
    for start in range(0, len(registros_finales), BATCH_SIZE):
        batch = registros_finales[start : start + BATCH_SIZE]
        await db.execute(insert(Cliente), batch)
        await db.flush()
        creados += len(batch)
        logger.info(f"Clientes importados: batch {start}–{start + len(batch)}")

    await db.commit()

    return {"creados": creados, "errores": errores}


async def importar_productos(db: AsyncSession, file_bytes: bytes) -> dict:
    """
    Importa productos desde un archivo Excel.
    
    Misma estrategia de performance que importar_clientes.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Archivo Excel inválido o corrupto")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene datos (las filas empiezan en la fila 3)")

    # ── Fase 1: Validación en memoria ──
    errores = []
    registros_validos = []
    codigos_en_archivo = set()
    tipos_validos = {"producto", "servicio", "combo"}
    monedas_validas = {"PEN", "USD"}

    for i, row in enumerate(rows, start=3):
        padded = list(row) + [None] * (11 - len(row)) if len(row) < 11 else list(row)

        codigo = _clean(padded[0])
        nombre = _clean(padded[1])
        precio_raw = padded[2]

        # Validar campos obligatorios
        if not codigo or not nombre or precio_raw is None:
            errores.append({"fila": i, "error": "Faltan campos obligatorios (codigo, nombre, precio_unitario)"})
            continue

        # Validar precio
        try:
            precio = float(precio_raw)
            if precio < 0:
                raise ValueError()
        except (ValueError, TypeError):
            errores.append({"fila": i, "error": f"precio_unitario inválido: '{precio_raw}'. Debe ser un número positivo"})
            continue

        # Tipo
        tipo = (_clean(padded[3]) or "producto").lower()
        if tipo not in tipos_validos:
            errores.append({"fila": i, "error": f"tipo inválido: '{tipo}'. Use: producto, servicio o combo"})
            continue

        # Moneda
        moneda = (_clean(padded[8]) or "PEN").upper()
        if moneda not in monedas_validas:
            errores.append({"fila": i, "error": f"moneda inválida: '{moneda}'. Use: PEN o USD"})
            continue

        # Costo unitario
        costo = None
        if padded[7] is not None:
            try:
                costo = float(padded[7])
            except (ValueError, TypeError):
                errores.append({"fila": i, "error": f"costo_unitario inválido: '{padded[7]}'"})
                continue

        # IGV
        igv = 18.0
        if padded[10] is not None:
            try:
                igv = float(padded[10])
            except (ValueError, TypeError):
                errores.append({"fila": i, "error": f"igv_porcentaje inválido: '{padded[10]}'"})
                continue

        # Duplicados en archivo
        if codigo in codigos_en_archivo:
            errores.append({"fila": i, "error": f"codigo '{codigo}' duplicado en el archivo"})
            continue
        codigos_en_archivo.add(codigo)

        registros_validos.append({
            "codigo": codigo,
            "nombre": nombre,
            "precio_unitario": precio,
            "tipo": tipo,
            "descripcion": _clean(padded[4]),
            "categoria": _clean(padded[5]),
            "marca": _clean(padded[6]),
            "costo_unitario": costo,
            "moneda": moneda,
            "unidad_medida": _clean(padded[9]) or "UND",
            "igv_porcentaje": igv,
            "aplica_igv": igv > 0,
            "tipo_afectacion_igv": "10" if igv > 0 else "20",
            "estado": "activo",
        })

    if not registros_validos:
        return {"creados": 0, "errores": errores}

    # ── Fase 2: Verificar códigos duplicados contra la BD ──
    codigos_a_insertar = [r["codigo"] for r in registros_validos]
    result = await db.execute(
        select(Producto.codigo).where(
            Producto.codigo.in_(codigos_a_insertar),
            Producto.deleted_at == None,
        )
    )
    codigos_existentes = set(result.scalars().all())

    registros_finales = []
    for r in registros_validos:
        if r["codigo"] in codigos_existentes:
            errores.append({"fila": "—", "error": f"codigo '{r['codigo']}' ya existe en la base de datos"})
        else:
            registros_finales.append(r)

    if not registros_finales:
        return {"creados": 0, "errores": errores}

    # ── Fase 3: INSERT por lotes ──
    creados = 0
    for start in range(0, len(registros_finales), BATCH_SIZE):
        batch = registros_finales[start : start + BATCH_SIZE]
        await db.execute(insert(Producto), batch)
        await db.flush()
        creados += len(batch)
        logger.info(f"Productos importados: batch {start}–{start + len(batch)}")

    await db.commit()

    return {"creados": creados, "errores": errores}
